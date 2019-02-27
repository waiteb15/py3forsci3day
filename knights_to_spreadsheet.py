#!/usr/bin/env python
"""
prints knights.txt to knights.xlsx

"""

import openpyxl as px

WB = px.Workbook()
WS = WB.active
WS.title = 'knights'
HEADING = "Name, Title, Favorite Color, Quest, Comment".split(',')

for i, val in enumerate(HEADING):
    WS.cell(row=1, column=i+1).value = val

with open('DATA/knights.txt') as knights_in:
    for ROWVAL, row in enumerate(knights_in):
        vals = row.rstrip('\n').split(':')
        for i, val in enumerate(vals):
            WS.cell(row=ROWVAL+2, column=i + 1).value = val

WB.save('knights.xlsx')
