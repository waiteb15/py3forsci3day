#!/usr/bin/env python
# (c)2015 John Strickler

import openpyxl as px

def main():
    """program entry point"""
    wb = px.load_workbook('../DATA/presidents.xlsx')
    ws = wb.get_sheet_by_name('US Presidents')

    print(ws.dimensions)
    print(ws.min_col)
    print(ws.min_row)
    print(ws.max_column)
    print(ws.max_row)
    # returns value for specified cell
    print(ws.cell(row=2, column=3).value, ws.cell(row=2, column=2).value, '\n')
    # ws.rows is all rows
    # ws.columns is all columns
    print(ws.rows[1][2].value, ws.rows[1][1].value, '\n')
    for i in range(1,6):
        print(ws.columns[1][i].value)
    print()

if __name__ == '__main__':
    main()
