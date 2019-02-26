#!/usr/bin/env python
# (c)2015 John Strickler

import openpyxl as px
print(px.__version__)  # should be >= 2.2 or this code will not work

def main():
    wb = px.load_workbook('../DATA/presidents.xlsx')

# three ways to get to a worksheet:

    # 1
    print(wb.get_sheet_names(), '\n')
    ws = wb.get_sheet_by_name('US Presidents')
    # or
    ws = wb['US Presidents']
    print(ws, '\n')

    # 2
    for ws in wb:
        print(ws)
    print()

    # 3
    ws = wb.active
    print(ws, '\n')

    print(ws['A1'].value)
    print(ws['C2'].value, ws['B2'].value)


if __name__ == '__main__':
    main()
